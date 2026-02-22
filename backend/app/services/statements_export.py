"""
Export extracted statements (and extraction summary) to CSV or Excel for tracking.
"""
from __future__ import annotations

import csv
import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook

# Characters illegal in Excel/OpenPyXL (control chars 0x00-0x08, 0x0B-0x0C, 0x0E-0x1F, 0x7F)
_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


def _sanitize_for_xlsx(val: Any) -> Any:
    """Remove illegal characters from strings; numbers pass through."""
    if val is None:
        return ""
    if isinstance(val, (int, float)):
        return val
    s = str(val)
    return _ILLEGAL_XLSX_RE.sub("", s)


from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _period_columns(periods_json: list) -> list[str]:
    """Extract period labels from Statement.periods_json for column headers."""
    if not periods_json:
        return []
    return [p.get("label", str(i)) for i, p in enumerate(periods_json)]


def _row_values(line: Any, period_labels: list[str]) -> list:
    """Values for a statement line in period order; values_json is {period_key: value}."""
    import re
    values_json = getattr(line, "values_json", None) or {}
    out: list = []
    for lbl in period_labels:
        v = values_json.get(lbl)
        if v is None and lbl:
            # Fallback: period label may differ (e.g. "2025" vs "52 weeks 2025 Rm"); match by year
            m = re.search(r"\b(20\d{2})\b", str(lbl))
            if m:
                yr = m.group(1)
                for k in values_json:
                    if yr in str(k) or k == yr:
                        v = values_json.get(k)
                        if v is not None:
                            break
        out.append(v if v is not None else "")
    return out


def _soce_column_labels() -> dict[str, str]:
    """Display labels for SOCE column keys."""
    return {
        "total_equity": "Total equity",
        "non_controlling_interest": "Non-controlling interest",
        "attributable_total": "Attributable total",
        "stated_capital": "Stated capital",
        "treasury_shares": "Treasury shares",
        "other_reserves": "Other reserves",
        "retained_earnings": "Retained earnings",
    }


# Label substrings/patterns that indicate a total or subtotal row (SFP, SCI, CF)
_TOTAL_LABEL_PATTERNS = (
    "total ",
    "balance at",
    "cash flows from operating",
    "cash flows utilised by",
    "net movement in cash",
    "cash and cash equivalents at the end",
    "cash generated from operations",
    "total assets",
    "total equity",
    "total liabilities",
    "total current assets",
    "total non-current assets",
    "total current liabilities",
    "total non-current liabilities",
    "total equity and liabilities",
    "profit for the year",
    "total comprehensive income",
    "profit/(loss) for the year",
    "comprehensive income for the year",
)


def _is_total_row(line: Any, raw_label: str) -> bool:
    """True if this line is a total/subtotal that should use =SUM() in Excel."""
    role = None
    ev = getattr(line, "evidence_json", None) or {}
    if isinstance(ev, dict):
        role = ev.get("row_role")
    if role in ("subtotal", "total"):
        return True
    lower = raw_label.strip().lower()
    return any(p in lower for p in _TOTAL_LABEL_PATTERNS)


def _row_values_soce(line: Any, periods_json: list) -> list:
    """Values for SoCE line: values_json is {period: {column: value}}."""
    import re
    values_json = getattr(line, "values_json", None) or {}
    out: list = []
    for p in periods_json:
        period_label = p.get("label", "")
        columns = p.get("columns") or []
        period_vals = values_json.get(period_label) if isinstance(values_json.get(period_label), dict) else {}
        if not period_vals and period_label:
            # Fallback: match by year (e.g. "52 weeks 2025" vs stored "2025")
            m = re.search(r"\b(20\d{2})\b", str(period_label))
            if m:
                yr = m.group(1)
                for k, v in values_json.items():
                    if isinstance(v, dict) and (k == yr or yr in str(k)):
                        period_vals = v
                        break
        for col in columns:
            v = period_vals.get(col) if period_vals else None
            out.append(v if v is not None and v != "" else "")
    return out


def build_statements_xlsx(
    version_id: str,
    document_filename: str,
    status: str,
    statements: list[Any],
    presentation_scale: dict | None = None,
    notes_index: list[Any] | None = None,
    note_extractions: list[Any] | None = None,
) -> BytesIO:
    """
    Build an Excel workbook:
    - Summary sheet: version, filename, status, scale/currency, step tracking.
    - One sheet per statement (SFP, SCI, CF, SoCE): line_no, raw_label, section_path, period columns, source_pages.
    - Notes Summary tab: note_no, title, start_page, end_page, confidence.
    - Per-type tabs: Borrowings, Leases, Contingencies, Risk (extracted fields).
    """
    wb = Workbook()

    # --- Summary sheet ---
    ws_summary = wb.active
    if ws_summary is None:
        raise RuntimeError("No active sheet")
    ws_summary.title = "Summary"
    ws_summary.append(["Extraction export – document version"])
    ws_summary.append([])
    ws_summary.append(["Document version ID", version_id])
    ws_summary.append(["Original filename", document_filename])
    ws_summary.append(["Status", status])
    if presentation_scale:
        ws_summary.append(["Currency", presentation_scale.get("currency") or ""])
        ws_summary.append(["Scale", presentation_scale.get("scale") or ""])
        ws_summary.append(["Scale factor", presentation_scale.get("scale_factor") or ""])
    ws_summary.append([])
    ws_summary.append(["Statements extracted", len(statements)])
    for s in statements:
        line_count = len(getattr(s, "lines", []))
        ws_summary.append([f"  {s.statement_type} ({s.entity_scope})", f"{line_count} lines"])
    if notes_index:
        ws_summary.append([])
        ws_summary.append(["Notes index entries", len(notes_index)])
    if note_extractions:
        ws_summary.append(["Note extractions", len(note_extractions)])

    # --- Notes Summary tab ---
    if notes_index:
        ws_notes = wb.create_sheet(title="Notes_Summary")
        ws_notes.append(["note_number", "title", "start_page", "end_page", "confidence"])
        for ni in notes_index:
            ws_notes.append([
                getattr(ni, "note_number", ""),
                _sanitize_for_xlsx((getattr(ni, "title", "") or "")[:200]),
                getattr(ni, "start_page", ""),
                getattr(ni, "end_page", ""),
                getattr(ni, "confidence", ""),
            ])

    # --- Per-type note tabs (group by type to avoid duplicate sheet names) ---
    if note_extractions:
        by_type: dict[str, list[tuple[Any, dict]]] = {}
        for ne in note_extractions:
            ev = getattr(ne, "evidence_json", None) or {}
            ext = ev.get("extraction", ev)
            ne_type = ext.get("type", "OTHER")
            fields = ext.get("fields", {})
            if not fields:
                continue
            by_type.setdefault(ne_type, []).append((ne, ext))
        for ne_type, items in by_type.items():
            sheet_name = f"Note_{ne_type}"[:31]
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["note_number", "field", "2025", "2024", "unit"])
            for ne, ext in items:
                note_num = getattr(ne, "note_number", "")
                fields = ext.get("fields", {})
                ev = getattr(ne, "evidence_json", None) or {}
                pages = ev.get("pages", [])
                for k, v in fields.items():
                    if k == "unit":
                        ws.append([note_num, "unit", _sanitize_for_xlsx(v), _sanitize_for_xlsx(v), _sanitize_for_xlsx(v)])
                    elif isinstance(v, dict):
                        y1 = v.get("2025", v.get("2024", ""))
                        y2 = v.get("2024", v.get("2025", ""))
                        ws.append([note_num, _sanitize_for_xlsx(k), _sanitize_for_xlsx(y1), _sanitize_for_xlsx(y2), ""])
                if pages:
                    ws.append([note_num, "source_pages", ",".join(str(p) for p in pages), "", ""])

    # --- Helper styles for statement sheets ---
    title_font = Font(size=14, bold=True, color="C00000")
    section_font = Font(size=11, bold=True)
    subsection_font = Font(size=11, bold=True, color="C00000")
    header_font = Font(size=11, bold=True)
    total_font = Font(size=11, bold=True)
    center_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_text = Alignment(horizontal="left", vertical="center")
    right_number = Alignment(horizontal="right", vertical="center")
    thin_side = Side(style="thin", color="CCCCCC")
    total_border = Border(top=thin_side)
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    total_fill = PatternFill("solid", fgColor="DDDDDD")

    statement_titles = {
        "SFP": "Consolidated statement of financial position",
        "SCI": "Consolidated statement of profit or loss and other comprehensive income",
        "SOCE": "Consolidated statement of comprehensive income",
        "CF": "Consolidated statement of cash flows",
        "SoCE": "Statement of changes in equity",
        "IS": "Income statement",
    }

    # --- One sheet per statement ---
    for stmt in statements:
        lines = list(getattr(stmt, "lines", []))
        periods_json = getattr(stmt, "periods_json", None) or []
        period_labels = _period_columns(periods_json)

        # SoCE: multi-column layout (Total equity, NCI, Stated capital, etc.) per period
        is_soce = stmt.statement_type == "SoCE" and periods_json and isinstance(periods_json[0].get("columns"), list)
        if is_soce:
            # Build headers: Line item | Notes | [value columns] | Source pages
            # Use header_labels verbatim when present (from LLM image extraction); else col_labels mapping
            col_labels = _soce_column_labels()
            value_headers: list[str] = []
            header_labels = periods_json[0].get("header_labels") if periods_json else []
            for p in periods_json:
                period_label = p.get("label", "")
                columns = p.get("columns") or []
                labels = p.get("header_labels") or header_labels
                for i, col in enumerate(columns):
                    if labels and i < len(labels):
                        label = labels[i]
                    else:
                        label = col_labels.get(col, col.replace("_", " ").title())
                    value_headers.append(f"{label} ({period_label})" if period_label else label)
            headers = ["Line item", "Notes"] + value_headers + ["Source pages"]
        else:
            headers = ["Line item", "Notes"] + period_labels + ["Source pages"]

        # Clean, predictable sheet names so SFP / SoCE are on clear separate tabs
        base_name = stmt.statement_type
        if getattr(stmt, "entity_scope", None) and stmt.entity_scope != "GROUP":
            base_name = f"{base_name}_{stmt.entity_scope}"
        sheet_name = base_name[:31]

        ws = wb.create_sheet(title=sheet_name)

        # Title row to mirror PDF heading
        title = statement_titles.get(stmt.statement_type, stmt.statement_type)
        if getattr(stmt, "entity_scope", None):
            title = f"{title} – {stmt.entity_scope.title()}"
        ws.append([_sanitize_for_xlsx(title)])
        max_col = len(headers)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = title_font
        title_cell.alignment = left_text

        # Blank spacer row
        ws.append([])

        # Header row
        header_row_idx = ws.max_row + 1
        ws.append([_sanitize_for_xlsx(h) for h in headers])
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.font = header_font
            cell.alignment = center_header if col_idx >= 3 else left_text
            cell.fill = header_fill

        # Set column widths
        ws.column_dimensions["A"].width = 55  # Line item
        ws.column_dimensions["B"].width = 10  # Notes
        num_value_cols = len(headers) - 3  # minus Line item, Notes, Source pages
        for i in range(num_value_cols):
            ws.column_dimensions[get_column_letter(3 + i)].width = 14
        ws.column_dimensions[get_column_letter(3 + num_value_cols)].width = 18  # Source pages

        # Data rows with section grouping similar to PDF layout
        seen_top_sections: set[str] = set()
        last_subsection_by_top: dict[str, str] = {}
        # For SFP/SCI/CF: track first data row of current block to build =SUM() formulas
        block_start_row: int | None = None

        for line in lines:
            raw_label = (getattr(line, "raw_label", "") or "")[:500]
            section_path = (getattr(line, "section_path", "") or "")[:200]
            parts = [p.strip() for p in section_path.split(">") if p.strip()] if section_path else []

            top = parts[0] if parts else None
            sub = parts[1] if len(parts) > 1 else None

            # Insert top-level section row (e.g. "Assets", "Equity", "Liabilities")
            if top and top not in seen_top_sections:
                seen_top_sections.add(top)
                ws.append([_sanitize_for_xlsx(top)])
                r = ws.max_row
                cell = ws.cell(row=r, column=1)
                cell.font = section_font
                cell.alignment = left_text

            # Insert subsection row (e.g. "Non-current assets", "Current assets")
            if top and sub and last_subsection_by_top.get(top) != sub:
                last_subsection_by_top[top] = sub
                ws.append([_sanitize_for_xlsx(sub)])
                r = ws.max_row
                cell = ws.cell(row=r, column=1)
                cell.font = subsection_font
                cell.alignment = left_text

            evidence = getattr(line, "evidence_json", None) or {}
            pages = evidence.get("pages", [])
            pages_str = ",".join(str(p) for p in pages) if pages else ""

            # Notes column from note_refs_json
            note_refs = getattr(line, "note_refs_json", None) or []
            if isinstance(note_refs, list):
                notes_str = ",".join(str(n) for n in note_refs if n)
            else:
                notes_str = str(note_refs) if note_refs else ""

            values = (
                _row_values_soce(line, periods_json)
                if is_soce
                else _row_values(line, period_labels)
            )

            row_idx = ws.max_row + 1
            num_val_cols = len(values)
            is_total = _is_total_row(line, raw_label)

            # For SFP/SCI/CF totals: use =SUM() formulas instead of values
            use_sum_formula = (
                not is_soce
                and is_total
                and block_start_row is not None
                and block_start_row <= row_idx - 1
            )
            if use_sum_formula:
                # Placeholder values; we'll overwrite with formulas after append
                row_values_for_append = [""] * num_val_cols
            else:
                row_values_for_append = [_sanitize_for_xlsx(v) for v in values]

            row_data = (
                [_sanitize_for_xlsx(raw_label), _sanitize_for_xlsx(notes_str)]
                + row_values_for_append
                + [_sanitize_for_xlsx(pages_str)]
            )
            ws.append(row_data)

            # Overwrite value cells with =SUM() for total rows (SFP/SCI/CF)
            if use_sum_formula:
                for i in range(num_val_cols):
                    col_idx = 3 + i
                    col_letter = get_column_letter(col_idx)
                    formula = f"=SUM({col_letter}{block_start_row}:{col_letter}{row_idx - 1})"
                    ws.cell(row=row_idx, column=col_idx).value = formula
            if not is_soce:
                if is_total:
                    block_start_row = None  # next block will set on first component
                elif block_start_row is None:
                    block_start_row = row_idx

            # Styling for the just-added data row
            depth = len(parts)

            # Line item cell
            cell_label = ws.cell(row=row_idx, column=1)
            cell_label.alignment = Alignment(horizontal="left", vertical="center", indent=max(depth - 1, 0))
            if is_total:
                cell_label.font = total_font

            # Notes cell
            ws.cell(row=row_idx, column=2).alignment = center_header

            # Value cells
            for i in range(num_val_cols):
                col_idx = 3 + i
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = right_number
                cell.number_format = '#,##0;[Red]-#,##0'
                if is_total:
                    cell.font = total_font

            # Source pages cell
            ws.cell(row=row_idx, column=3 + num_val_cols).alignment = center_header

            # Highlight totals with a top border and light fill
            if is_total:
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = total_border
                    cell.fill = total_fill

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_statements_csv(
    version_id: str,
    document_filename: str,
    status: str,
    statements: list[Any],
    presentation_scale: dict | None = None,
    notes_index: list[Any] | None = None,
    note_extractions: list[Any] | None = None,
) -> BytesIO:
    """
    Single CSV with statement_type so all steps can be tracked in one file.
    Columns: statement_type, entity_scope, line_no, raw_label, section_path, period_1, period_2, ..., source_pages.
    """
    buf = BytesIO()
    writer = csv.writer(buf)

    # Header comment rows for tracking
    writer.writerow(["# Extraction export", version_id])
    writer.writerow(["# Filename", document_filename])
    writer.writerow(["# Status", status])
    if presentation_scale:
        writer.writerow(["# Currency", presentation_scale.get("currency") or ""])
        writer.writerow(["# Scale", presentation_scale.get("scale") or ""])
    if notes_index:
        writer.writerow(["# Notes index entries", len(notes_index)])
    if note_extractions:
        writer.writerow(["# Note extractions", len(note_extractions)])
    writer.writerow([])

    # Notes Summary (if present)
    if notes_index:
        writer.writerow(["# Notes_Summary"])
        writer.writerow(["note_number", "title", "start_page", "end_page", "confidence"])
        for ni in notes_index:
            writer.writerow([
                getattr(ni, "note_number", ""),
                (getattr(ni, "title", "") or "")[:200],
                getattr(ni, "start_page", ""),
                getattr(ni, "end_page", ""),
                getattr(ni, "confidence", ""),
            ])
        writer.writerow([])

    # Collect value column headers: use SoCE multi-column if any stmt is SoCE, else period labels
    has_soce = any(
        stmt.statement_type == "SoCE"
        and getattr(stmt, "periods_json", None)
        and isinstance((getattr(stmt, "periods_json") or [{}])[0].get("columns"), list)
        for stmt in statements
    )
    all_value_headers: list[str] = []
    if has_soce:
        for stmt in statements:
            if stmt.statement_type != "SoCE":
                continue
            periods_json = getattr(stmt, "periods_json", None) or []
            if not periods_json or not isinstance(periods_json[0].get("columns"), list):
                continue
            col_labels = _soce_column_labels()
            for p in periods_json:
                period_label = p.get("label", "")
                for col in p.get("columns") or []:
                    label = col_labels.get(col, col.replace("_", " "))
                    all_value_headers.append(f"{label} ({period_label})" if period_label else label)
            break
    if not all_value_headers:
        for stmt in statements:
            p = _period_columns(getattr(stmt, "periods_json", None) or [])
            for lbl in p:
                if lbl not in all_value_headers:
                    all_value_headers.append(lbl)
    if not all_value_headers:
        all_value_headers = ["period_1", "period_2"]

    headers = ["statement_type", "entity_scope", "line_no", "raw_label", "section_path"] + all_value_headers + ["source_pages"]
    writer.writerow(headers)

    for stmt in statements:
        periods_json = getattr(stmt, "periods_json", None) or []
        is_soce = stmt.statement_type == "SoCE" and periods_json and isinstance(periods_json[0].get("columns"), list)
        period_labels = _period_columns(periods_json)
        for line in getattr(stmt, "lines", []):
            evidence = getattr(line, "evidence_json", None) or {}
            pages = evidence.get("pages", [])
            pages_str = ",".join(str(p) for p in pages) if pages else ""
            values = (
                _row_values_soce(line, periods_json)
                if is_soce
                else _row_values(line, period_labels)
            )
            while len(values) < len(all_value_headers):
                values.append("")
            row = [
                stmt.statement_type,
                stmt.entity_scope,
                getattr(line, "line_no", ""),
                (getattr(line, "raw_label", "") or "")[:500],
                (getattr(line, "section_path", "") or "")[:200],
            ] + values[: len(all_value_headers)] + [pages_str]
            writer.writerow(row)

    buf.seek(0)
    return buf


def build_mappings_csv(
    version_id: str,
    document_filename: str,
    status: str,
    canonical_mappings: dict | None,
) -> BytesIO:
    """
    Export canonical mappings (raw_label -> canonical_key) as CSV when Statement rows are empty.
    Columns: statement_type, section_path, raw_label, canonical_key, confidence, reason.
    """
    buf = BytesIO()
    writer = csv.writer(buf)
    writer.writerow(["# Canonical mappings export", version_id])
    writer.writerow(["# Filename", document_filename])
    writer.writerow(["# Status", status])
    writer.writerow([])
    writer.writerow(["statement_type", "section_path", "raw_label", "canonical_key", "confidence", "reason"])
    mappings = (canonical_mappings or {}).get("mappings") or []
    for m in mappings:
        section_path = m.get("section_path")
        if isinstance(section_path, list):
            section_path = " > ".join(str(x) for x in section_path)
        writer.writerow([
            m.get("statement_type", ""),
            section_path or "",
            (m.get("raw_label") or "")[:500],
            m.get("canonical_key", "UNMAPPED"),
            m.get("confidence", ""),
            (m.get("reason") or "")[:300],
        ])
    buf.seek(0)
    return buf


def build_mappings_xlsx(
    version_id: str,
    document_filename: str,
    status: str,
    canonical_mappings: dict | None,
) -> BytesIO:
    """Export canonical mappings as Excel when Statement rows are empty.

    Layout:
    - Summary sheet with high-level metadata.
    - One sheet per statement_type (e.g. SFP, SCI, CF, SoCE) so that
      SFP and SoCE each live on their own tab.
    """
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("No active sheet")
    ws.title = "Summary"
    ws.append(["Canonical mappings export"])
    ws.append([])
    ws.append(["Document version ID", version_id])
    ws.append(["Filename", document_filename])
    ws.append(["Status", status])
    ws.append([])

    # Group mappings by statement_type so SFP / SoCE each get their own tab
    mappings = (canonical_mappings or {}).get("mappings") or []
    by_type: dict[str, list[dict]] = {}
    for m in mappings:
        stmt_type = (m.get("statement_type") or "OTHER").strip() or "OTHER"
        by_type.setdefault(stmt_type, []).append(m)

    # Deterministic sheet ordering
    for stmt_type in sorted(by_type.keys()):
        # Keep sheet names short and Excel-safe
        sheet_name = stmt_type[:31]
        ws_mappings = wb.create_sheet(title=sheet_name)
        # Header row – we still export the same fields, but
        # on a per-statement tab so it's easier to work with.
        ws_mappings.append(
            ["section_path", "raw_label", "canonical_key", "confidence", "reason"]
        )

        for m in by_type[stmt_type]:
            section_path = m.get("section_path")
            if isinstance(section_path, list):
                section_path = " > ".join(str(x) for x in section_path)
            ws_mappings.append(
                [
                    _sanitize_for_xlsx(section_path or ""),
                    _sanitize_for_xlsx((m.get("raw_label") or "")[:500]),
                    _sanitize_for_xlsx(m.get("canonical_key", "UNMAPPED")),
                    _sanitize_for_xlsx(m.get("confidence", "")),
                    _sanitize_for_xlsx((m.get("reason") or "")[:300]),
                ]
            )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
