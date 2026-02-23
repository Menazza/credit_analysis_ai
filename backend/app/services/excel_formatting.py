"""
Excel formatting for financial statement extraction output.
Shared by worker pipeline and test script.
"""
from __future__ import annotations

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def format_statement_sheet(worksheet, df, statement_type: str) -> None:
    """Apply formatting to a statement sheet (bold totals, blue header, column widths)."""
    bold_font = Font(bold=True)
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    thin_border = Side(style="thin", color="000000")
    top_bottom_border = Border(top=thin_border, bottom=thin_border)

    total_patterns = {
        "SFP": ["total assets", "total equity", "total liabilities", "total current", "total non-current"],
        "SCI": ["total comprehensive", "profit for the", "revenue", "operating profit", "gross profit"],
        "CF": ["cash generated", "net increase", "net decrease", "cash and cash equivalents at"],
        "SOCE": ["balance at", "total comprehensive", "dividends"],
    }
    patterns = total_patterns.get(statement_type, ["total"])

    label_col_idx = list(df.columns).index("raw_label") + 1 if "raw_label" in df.columns else None

    for row_idx, row in df.iterrows():
        excel_row = row_idx + 2
        raw_label = str(row.get("raw_label", "")).lower()
        section = str(row.get("section", "")) if row.get("section") else ""

        is_total_row = any(p in raw_label for p in patterns)
        max_col = len(df.columns)

        if is_total_row or "balance at" in raw_label:
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=excel_row, column=col)
                cell.font = bold_font
                cell.fill = grey_fill
                cell.border = top_bottom_border
        elif section and label_col_idx:
            cell = worksheet.cell(row=excel_row, column=label_col_idx)
            cell.alignment = Alignment(indent=1)

    # Header row: blue background, white text
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    bottom_border = Border(bottom=thin_border)

    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = bottom_border

    # Auto-adjust column widths
    for idx, col in enumerate(df.columns):
        try:
            max_len = max(
                df[col].astype(str).str.len().max() if len(df) > 0 else 0,
                len(str(col)),
            ) + 2
        except Exception:
            max_len = len(str(col)) + 2
        col_letter = get_column_letter(idx + 1)
        worksheet.column_dimensions[col_letter].width = min(max_len, 50)

    worksheet.freeze_panes = "C2"


def format_summary_sheet(worksheet) -> None:
    """Apply formatting to the Summary sheet."""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, 3):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font

    worksheet.column_dimensions["A"].width = 20
    worksheet.column_dimensions["B"].width = 40
