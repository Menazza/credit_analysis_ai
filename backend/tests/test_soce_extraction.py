"""
Test script for SoCE extraction using the same pipeline as the worker.
Processes a PDF and outputs the extracted data to an Excel file for inspection.

Usage:
    cd backend
    python -m tests.test_soce_extraction
    
    # Or specify a specific page:
    python -m tests.test_soce_extraction --page 22
"""
import os
import sys
import argparse
from pathlib import Path

# Add backend to path
backend_dir = Path(__file__).parent.parent
sys.path.insert(0, str(backend_dir))

import pandas as pd
from datetime import datetime


def find_soce_pages_with_llm(pdf_bytes: bytes) -> list[int]:
    """Use LLM to find SoCE pages like the worker does."""
    from app.services.llm.tasks import llm_classify_statement_regions
    import fitz
    
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    regions = []
    
    # Build regions from pages (similar to what worker does)
    for i in range(len(doc)):
        page = doc[i]
        text = page.get_text()
        if text.strip():
            regions.append({
                "region_id": f"page_{i+1}",
                "page": i + 1,
                "text": text[:3000]  # First 3000 chars for classification
            })
    
    doc.close()
    
    # Call LLM to classify
    print("Calling LLM to classify pages...")
    try:
        result = llm_classify_statement_regions(regions, doc_hash=None)
        if result and result.regions:
            soce_pages = [
                int(r.region_id.replace("page_", ""))
                for r in result.regions
                if r.statement_type == "SOCE"
            ]
            print(f"LLM identified SoCE pages: {soce_pages}")
            return soce_pages
    except Exception as e:
        print(f"LLM classification failed: {e}")
    
    return []


def find_soce_pages_heuristic(pdf_bytes: bytes) -> list[int]:
    """Fallback heuristic to find SoCE pages."""
    import fitz
    
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    soce_pages = []
    
    for i in range(len(doc)):
        text = doc[i].get_text().lower()
        # Look for SoCE indicators
        has_title = "statement of changes in equity" in text or "changes in equity" in text
        has_columns = "total equity" in text and ("retained" in text or "treasury" in text or "stated capital" in text)
        has_balance = "balance at" in text
        
        if (has_title and has_columns) or (has_columns and has_balance):
            soce_pages.append(i + 1)
    
    doc.close()
    return soce_pages


def test_soce_extraction(pdf_path: str, output_dir: str = None, specific_page: int = None, use_llm: bool = True):
    """
    Extract SoCE from a PDF and save results to Excel.
    
    Args:
        pdf_path: Path to the PDF file
        output_dir: Directory to save results (default: test_results in project root)
        specific_page: If provided, only test this page number
        use_llm: Whether to use LLM for page classification (default: True)
    """
    from app.services.soce_geometry_extractor import extract_soce_structured_lines_geometry
    
    # Setup output directory
    if output_dir is None:
        output_dir = backend_dir.parent / "test_results"
    else:
        output_dir = Path(output_dir)
    
    output_dir.mkdir(exist_ok=True)
    
    # Read PDF
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        print(f"ERROR: PDF not found at {pdf_path}")
        return
    
    print(f"Reading PDF: {pdf_path}")
    with open(pdf_path, "rb") as f:
        pdf_bytes = f.read()
    
    # Find SoCE pages
    if specific_page:
        soce_pages = [specific_page]
        print(f"Testing specific page: {specific_page}")
    elif use_llm:
        soce_pages = find_soce_pages_with_llm(pdf_bytes)
        if not soce_pages:
            print("LLM found no SoCE pages, trying heuristic...")
            soce_pages = find_soce_pages_heuristic(pdf_bytes)
    else:
        soce_pages = find_soce_pages_heuristic(pdf_bytes)
    
    if not soce_pages:
        print("ERROR: Could not find any SoCE pages")
        return
    
    print(f"Testing SoCE extraction on pages: {soce_pages}")
    
    all_results = []
    all_column_headers = []
    
    for page_no in soce_pages:
        print(f"\n--- Page {page_no} ---")
        try:
            # First try geometry extractor
            lines = extract_soce_structured_lines_geometry(pdf_bytes, page_no, start_line_no=1)
            
            if not lines:
                print(f"  Geometry extractor found nothing, trying LLM...")
                # Fallback to LLM
                from app.services.soce_page_image import render_pdf_page_to_png
                from app.services.llm.tasks import llm_soce_table_from_image
                import base64
                
                png_bytes = render_pdf_page_to_png(pdf_bytes, page_no)
                b64 = base64.b64encode(png_bytes).decode("ascii")
                
                import fitz
                doc = fitz.open(stream=pdf_bytes, filetype="pdf")
                page_text = doc[page_no - 1].get_text()
                doc.close()
                
                out = llm_soce_table_from_image(b64, "test", page_no, doc_hash=None, pdf_text=page_text)
                
                if out and out.lines and out.column_headers:
                    headers = list(out.column_headers)
                    columns = [str(i) for i in range(len(headers))]
                    lines = []
                    for i, line in enumerate(out.lines):
                        vals = line.values if hasattr(line, "values") else []
                        values_by_col = {}
                        for j in range(len(columns)):
                            v = vals[j] if j < len(vals) else None
                            values_by_col[str(j)] = float(v) if v is not None else None
                        lines.append({
                            "line_no": 1 + i,
                            "raw_label": line.raw_label,
                            "note": line.note_ref,
                            "values_json": {"": values_by_col},
                            "section_path": line.section_path,
                            "column_keys": columns,
                            "column_headers": headers,
                            "period_labels": [""],
                        })
            
            if not lines:
                print(f"  No SoCE data found on page {page_no}")
                continue
            
            print(f"  Found {len(lines)} rows")
            
            # Get column info from first line
            column_keys = lines[0].get("column_keys", [])
            column_headers = lines[0].get("column_headers", [])
            
            print(f"  Column keys: {column_keys}")
            print(f"  Column headers: {column_headers}")
            
            if column_headers and column_headers not in all_column_headers:
                all_column_headers.append(column_headers)
            
            # Build rows for DataFrame
            for line in lines:
                row_data = {
                    "page": page_no,
                    "line_no": line.get("line_no"),
                    "raw_label": line.get("raw_label", ""),
                    "note": line.get("note"),
                    "section": line.get("section_path"),
                }
                
                # Extract values
                values_json = line.get("values_json", {})
                for period, cols in values_json.items():
                    for col_key, value in cols.items():
                        # Use header if available, otherwise key
                        try:
                            col_idx = int(col_key) if col_key.isdigit() else column_keys.index(col_key) if col_key in column_keys else -1
                            if col_idx >= 0 and col_idx < len(column_headers):
                                header = column_headers[col_idx]
                            else:
                                header = col_key
                        except (ValueError, IndexError):
                            header = col_key
                        
                        col_name = f"{header}" if not period else f"{header} ({period})"
                        row_data[col_name] = value
                
                all_results.append(row_data)
                
        except Exception as e:
            print(f"  Error on page {page_no}: {e}")
            import traceback
            traceback.print_exc()
    
    if not all_results:
        print("\nNo SoCE data extracted from any page.")
        return
    
    # Create DataFrame and save to Excel
    df = pd.DataFrame(all_results)
    
    # Generate output filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    pdf_name = pdf_path.stem
    output_file = output_dir / f"soce_test_{pdf_name}_{timestamp}.xlsx"
    
    # Save to Excel with formatting like the PDF
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="SoCE Extraction", index=False)
        
        worksheet = writer.sheets["SoCE Extraction"]
        
        # Define styles
        bold_font = Font(bold=True)
        grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        light_grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        thin_border = Side(style="thin", color="000000")
        top_bottom_border = Border(top=thin_border, bottom=thin_border)
        bottom_border = Border(bottom=thin_border)
        maroon_font = Font(color="800000", bold=True)
        
        # Get column indices
        label_col_idx = list(df.columns).index("raw_label") + 1 if "raw_label" in df.columns else None
        
        # Format each row based on its content
        for row_idx, row in df.iterrows():
            excel_row = row_idx + 2  # +2 because Excel is 1-indexed and has header row
            raw_label = str(row.get("raw_label", "")).lower()
            section = str(row.get("section", "")).lower() if row.get("section") else ""
            
            # Determine row type and apply formatting
            is_balance_row = "balance at" in raw_label
            is_total_row = "total comprehensive income" in raw_label
            is_dividends_row = "dividends" in raw_label
            is_section_header = is_total_row  # Total comprehensive income is a section header with values
            is_oci_item = "recognised in other comprehensive" in section
            
            # Apply formatting to all cells in the row
            max_col = len(df.columns)
            
            if is_balance_row:
                # Balance rows: bold, grey fill, top & bottom borders
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=excel_row, column=col)
                    cell.font = bold_font
                    cell.fill = grey_fill
                    cell.border = top_bottom_border
                    
            elif is_total_row:
                # Total comprehensive income: bold, light grey fill, top border
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=excel_row, column=col)
                    cell.font = bold_font
                    cell.fill = light_grey_fill
                    cell.border = Border(top=thin_border)
                    
            elif is_dividends_row:
                # Dividends: bold, light grey, bottom border
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=excel_row, column=col)
                    cell.font = bold_font
                    cell.fill = light_grey_fill
                    cell.border = bottom_border
                    
            elif is_oci_item:
                # OCI items (under "Recognised in other comprehensive loss"): indent label
                if label_col_idx:
                    cell = worksheet.cell(row=excel_row, column=label_col_idx)
                    cell.alignment = Alignment(indent=2)
            
            elif section == "total comprehensive income":
                # Items under Total comprehensive income (like Profit/loss): slight indent
                if label_col_idx:
                    cell = worksheet.cell(row=excel_row, column=label_col_idx)
                    cell.alignment = Alignment(indent=1)
        
        # Format header row
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = bold_font
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.border = Border(bottom=thin_border)
        
        # Auto-adjust column widths
        for idx, col in enumerate(df.columns):
            try:
                max_len = max(
                    df[col].astype(str).str.len().max() if len(df) > 0 else 0,
                    len(str(col))
                ) + 2
            except:
                max_len = len(str(col)) + 2
            col_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[col_letter].width = min(max_len, 50)
        
        # Freeze the header row and first few columns
        worksheet.freeze_panes = "D2"
    
    print(f"\n{'='*60}")
    print(f"RESULTS SAVED TO: {output_file}")
    print(f"{'='*60}")
    print(f"Total rows extracted: {len(all_results)}")
    
    # Also print a summary table
    print("\n=== EXTRACTION SUMMARY ===")
    value_cols = [c for c in df.columns if c not in ["page", "line_no", "raw_label", "note", "section"]]
    
    print(f"\nColumn headers detected: {all_column_headers}")
    print(f"\nValue columns in output: {value_cols}")
    print(f"\nAll rows:")
    display_cols = ["raw_label", "note", "section"] + value_cols[:7]
    display_cols = [c for c in display_cols if c in df.columns]
    pd.set_option('display.max_rows', 50)
    pd.set_option('display.max_colwidth', 50)
    print(df[display_cols].to_string())
    
    return output_file


def main():
    parser = argparse.ArgumentParser(description="Test SoCE extraction from PDF")
    parser.add_argument("pdf", nargs="?", help="Path to PDF file")
    parser.add_argument("--page", type=int, help="Test specific page number")
    parser.add_argument("--no-llm", action="store_true", help="Skip LLM classification, use heuristic only")
    
    args = parser.parse_args()
    
    # Default PDF path
    project_root = backend_dir.parent
    default_pdf = project_root / "shp-afs-2025.pdf"
    
    pdf_path = args.pdf if args.pdf else default_pdf
    
    print("=" * 60)
    print("SoCE EXTRACTION TEST")
    print("=" * 60)
    
    test_soce_extraction(
        pdf_path, 
        specific_page=args.page,
        use_llm=not args.no_llm
    )


if __name__ == "__main__":
    main()
