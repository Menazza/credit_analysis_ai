"""
Test script for extracting ALL financial statements from a PDF.
Uses the same detection pipeline as the worker, then parses each statement
and outputs to a formatted Excel file with one sheet per statement type.

Supported statements:
- SFP: Statement of Financial Position (Balance Sheet)
- SCI: Statement of Comprehensive Income (Income Statement)
- CF: Statement of Cash Flows
- SOCE: Statement of Changes in Equity

Usage:
    cd backend
    python tests/test_all_statements.py
    python tests/test_all_statements.py path/to/pdf.pdf
    python tests/test_all_statements.py --no-llm  # Skip LLM, use heuristic only
"""
import os
import sys
import argparse
import re
from pathlib import Path
from dataclasses import dataclass
from typing import Any

# Add backend to path
backend_dir = Path(__file__).parent.parent
sys.path.insert(0, str(backend_dir))

import pandas as pd
from datetime import datetime


def clean_for_excel(s):
    """Remove illegal characters for Excel."""
    if not isinstance(s, str):
        return s
    # Remove control characters that Excel doesn't allow
    import re
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', s)


@dataclass
class StatementPage:
    """A detected statement page."""
    page_no: int
    statement_type: str  # SFP, SCI, CF, SOCE
    entity_scope: str  # GROUP, COMPANY
    confidence: float


def classify_pages_with_llm(pdf_bytes: bytes) -> list[StatementPage]:
    """Use LLM to classify all pages by statement type."""
    from app.services.llm.tasks import llm_region_classifier
    import fitz
    
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    regions = []
    
    for i in range(len(doc)):
        page = doc[i]
        text = page.get_text()
        if text.strip():
            regions.append({
                "region_id": f"page_{i+1}",
                "page": i + 1,
                "text": text[:3000]
            })
    
    doc.close()
    
    print("Calling LLM to classify all pages...")
    try:
        result = llm_region_classifier(regions, document_version_id="test")
        if result and result.regions:
            pages = []
            for r in result.regions:
                if r.statement_type in ("SFP", "SCI", "IS", "CF", "SOCE"):
                    page_no = int(r.region_id.replace("page_", ""))
                    pages.append(StatementPage(
                        page_no=page_no,
                        statement_type=str(r.statement_type),
                        entity_scope=str(r.entity_scope),
                        confidence=r.confidence
                    ))
            print(f"LLM identified {len(pages)} statement pages")
            return pages
    except Exception as e:
        print(f"LLM classification failed: {e}")
        import traceback
        traceback.print_exc()
    
    return []


def classify_pages_heuristic(pdf_bytes: bytes) -> list[StatementPage]:
    """
    Heuristic to find PRIMARY statement pages.
    Handles two-column layouts where two statements appear on one page.
    """
    import fitz
    
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    pages = []
    
    for i in range(len(doc)):
        text = doc[i].get_text()
        text_lower = text.lower()
        page_no = i + 1
        
        # Skip pages that are clearly notes
        if "notes to the" in text_lower[:300]:
            continue
        
        # Skip index pages
        if page_no < 10:
            continue
        
        # Determine entity scope
        entity_scope = "GROUP"
        if "separate statement" in text_lower[:1000]:
            entity_scope = "COMPANY"
        elif "consolidated" in text_lower[:1000]:
            entity_scope = "GROUP"
        
        # Check for each statement type
        # This PDF has two-column layout, so multiple statements can be on one page
        
        # SFP detection
        if "statement of financial position" in text_lower[:1000]:
            if "total assets" in text_lower:
                pages.append(StatementPage(page_no, "SFP", entity_scope, 0.9))
        
        # SCI detection
        if "statement of comprehensive income" in text_lower[:1000]:
            if "revenue" in text_lower or "profit" in text_lower:
                pages.append(StatementPage(page_no, "SCI", entity_scope, 0.9))
        
        # SOCE detection
        if "statement of changes in equity" in text_lower[:1000]:
            if "balance at" in text_lower:
                pages.append(StatementPage(page_no, "SOCE", entity_scope, 0.9))
        
        # CF detection - note that CF might be on same page as SOCE (page 12)
        if "statement of cash flows" in text_lower[:1500]:
            if "operating" in text_lower or "cash generated" in text_lower or "cash flows" in text_lower:
                pages.append(StatementPage(page_no, "CF", entity_scope, 0.9))
    
    doc.close()
    
    # Deduplicate - keep PRIMARY statement page (skip index/contents pages)
    # Page 10 is typically the contents page, actual statements start at page 11
    seen = set()
    unique_pages = []
    
    # Sort pages - prefer pages 11-13 for GROUP statements (main statement pages)
    pages.sort(key=lambda p: (0 if 11 <= p.page_no <= 13 else 1, p.page_no))
    
    for p in pages:
        key = f"{p.statement_type}_{p.entity_scope}"
        if key not in seen:
            seen.add(key)
            unique_pages.append(p)
    
    return unique_pages


def extract_statement(pdf_bytes: bytes, page_no: int, statement_type: str) -> list[dict]:
    """
    Extract SFP, SCI, or CF statement using geometry extractor.
    Returns list of rows with raw_label, note, section, and period values.
    """
    from app.services.statement_geometry_extractor import extract_statement_structured_lines
    
    lines = extract_statement_structured_lines(pdf_bytes, page_no, statement_type)
    return lines


def extract_soce(pdf_bytes: bytes, page_no: int) -> list[dict]:
    """Extract SoCE using the geometry extractor."""
    from app.services.soce_geometry_extractor import extract_soce_structured_lines_geometry
    
    lines = extract_soce_structured_lines_geometry(pdf_bytes, page_no, start_line_no=1)
    return lines


def format_statement_sheet(worksheet, df, statement_type: str):
    """Apply formatting to a statement sheet like the PDF."""
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    
    bold_font = Font(bold=True)
    grey_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    light_grey_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    thin_border = Side(style="thin", color="000000")
    top_bottom_border = Border(top=thin_border, bottom=thin_border)
    bottom_border = Border(bottom=thin_border)
    
    # Define totals patterns based on statement type
    total_patterns = {
        "SFP": ["total assets", "total equity", "total liabilities", "total current", "total non-current"],
        "SCI": ["total comprehensive", "profit for the", "revenue", "operating profit", "gross profit"],
        "CF": ["cash generated", "net increase", "net decrease", "cash and cash equivalents at"],
        "SOCE": ["balance at", "total comprehensive", "dividends"],
    }
    patterns = total_patterns.get(statement_type, ["total"])
    
    label_col_idx = list(df.columns).index("raw_label") + 1 if "raw_label" in df.columns else None
    
    for row_idx, row in df.iterrows():
        excel_row = row_idx + 2
        raw_label = str(row.get("raw_label", "")).lower()
        section = str(row.get("section", "")).lower() if row.get("section") else ""
        
        is_total_row = any(p in raw_label for p in patterns)
        is_section_header = raw_label.upper() == raw_label and len(raw_label) < 50
        
        max_col = len(df.columns)
        
        if is_total_row or "balance at" in raw_label:
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=excel_row, column=col)
                cell.font = bold_font
                cell.fill = grey_fill
                cell.border = top_bottom_border
        elif section and label_col_idx:
            cell = worksheet.cell(row=excel_row, column=label_col_idx)
            cell.alignment = Alignment(indent=1)
    
    # Format header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = Border(bottom=thin_border)
    
    # Auto-adjust column widths
    for idx, col in enumerate(df.columns):
        try:
            max_len = max(
                df[col].astype(str).str.len().max() if len(df) > 0 else 0,
                len(str(col))
            ) + 2
        except:
            max_len = len(str(col)) + 2
        col_letter = get_column_letter(idx + 1)
        worksheet.column_dimensions[col_letter].width = min(max_len, 50)
    
    worksheet.freeze_panes = "C2"


def test_all_statements(pdf_path: str, output_dir: str = None, use_llm: bool = True):
    """
    Extract all financial statements from a PDF and save to Excel.
    Each statement type gets its own sheet.
    """
    if output_dir is None:
        output_dir = backend_dir.parent / "test_results"
    else:
        output_dir = Path(output_dir)
    
    output_dir.mkdir(exist_ok=True)
    
    pdf_path = Path(pdf_path)
    if not pdf_path.exists():
        print(f"ERROR: PDF not found at {pdf_path}")
        return
    
    print(f"Reading PDF: {pdf_path}")
    with open(pdf_path, "rb") as f:
        pdf_bytes = f.read()
    
    # Classify all pages
    if use_llm:
        statement_pages = classify_pages_with_llm(pdf_bytes)
        if not statement_pages:
            print("LLM found no statements, trying heuristic...")
            statement_pages = classify_pages_heuristic(pdf_bytes)
    else:
        statement_pages = classify_pages_heuristic(pdf_bytes)
    
    if not statement_pages:
        print("ERROR: Could not find any statement pages")
        return
    
    # Handle two-column pages: check for companion statements on the same page
    # Page 11: SFP (left) + SCI (right)
    # Page 12: SOCE (left) + CF (right)
    import fitz
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    additional_pages = []
    
    for sp in statement_pages:
        page_text = doc[sp.page_no - 1].get_text().lower()
        
        # If SFP is detected on a page, check for SCI
        if sp.statement_type == "SFP":
            if "statement of comprehensive income" in page_text or "profit for the year" in page_text:
                # Check if SCI is already detected for this page
                existing = [p for p in statement_pages if p.page_no == sp.page_no and p.statement_type == "SCI"]
                if not existing:
                    additional_pages.append(StatementPage(sp.page_no, "SCI", sp.entity_scope, 0.85))
        
        # If SOCE is detected on a page, check for CF
        if sp.statement_type == "SOCE":
            if "cash flows" in page_text or "operating activities" in page_text:
                existing = [p for p in statement_pages if p.page_no == sp.page_no and p.statement_type == "CF"]
                if not existing:
                    additional_pages.append(StatementPage(sp.page_no, "CF", sp.entity_scope, 0.85))
        
        # If CF is detected on a page, check for SOCE
        if sp.statement_type == "CF":
            if "statement of changes in equity" in page_text or "balance at" in page_text:
                existing = [p for p in statement_pages if p.page_no == sp.page_no and p.statement_type == "SOCE"]
                if not existing:
                    additional_pages.append(StatementPage(sp.page_no, "SOCE", sp.entity_scope, 0.85))
        
        # If SCI is detected on a page, check for SFP
        if sp.statement_type == "SCI":
            if "statement of financial position" in page_text or "total assets" in page_text:
                existing = [p for p in statement_pages if p.page_no == sp.page_no and p.statement_type == "SFP"]
                if not existing:
                    additional_pages.append(StatementPage(sp.page_no, "SFP", sp.entity_scope, 0.85))
    
    doc.close()
    statement_pages.extend(additional_pages)
    if additional_pages:
        print(f"Detected {len(additional_pages)} additional statements on two-column pages")
    
    # Group pages by statement type
    pages_by_type: dict[str, list[StatementPage]] = {}
    for sp in statement_pages:
        key = f"{sp.statement_type}_{sp.entity_scope}"
        if key not in pages_by_type:
            pages_by_type[key] = []
        pages_by_type[key].append(sp)
    
    print(f"\nDetected statements:")
    for key, pages in pages_by_type.items():
        print(f"  {key}: pages {[p.page_no for p in pages]}")
    
    # Extract each statement type
    all_dfs: dict[str, pd.DataFrame] = {}
    
    for key, pages in pages_by_type.items():
        statement_type = pages[0].statement_type
        entity_scope = pages[0].entity_scope
        
        print(f"\n--- Extracting {key} ---")
        all_rows = []
        all_column_headers = []
        
        for sp in pages:
            print(f"  Processing page {sp.page_no}...")
            
            try:
                if statement_type == "SOCE":
                    # Use SoCE geometry extractor
                    lines = extract_soce(pdf_bytes, sp.page_no)
                    
                    if lines:
                        column_headers = lines[0].get("column_headers", [])
                        if column_headers and column_headers not in all_column_headers:
                            all_column_headers.append(column_headers)
                        
                        for line in lines:
                            row_data = {
                                "page": sp.page_no,
                                "line_no": line.get("line_no"),
                                "raw_label": line.get("raw_label", ""),
                                "note": line.get("note"),
                                "section": line.get("section_path"),
                            }
                            
                            values_json = line.get("values_json", {})
                            col_headers = line.get("column_headers", [])
                            for period, cols in values_json.items():
                                for col_key, value in cols.items():
                                    try:
                                        col_idx = int(col_key) if col_key.isdigit() else -1
                                        if 0 <= col_idx < len(col_headers):
                                            header = col_headers[col_idx]
                                        else:
                                            header = col_key
                                    except (ValueError, IndexError):
                                        header = col_key
                                    
                                    col_name = f"{header}" if not period else f"{header} ({period})"
                                    row_data[col_name] = value
                            
                            all_rows.append(row_data)
                else:
                    # Use geometry extractor for SFP, SCI, CF
                    lines = extract_statement(pdf_bytes, sp.page_no, statement_type)
                    
                    if lines:
                        period_labels = lines[0].get("period_labels", [])
                        if period_labels and period_labels not in all_column_headers:
                            all_column_headers.append(period_labels)
                    
                    for line in lines:
                        row_data = {
                            "page": sp.page_no,
                            "line_no": line.get("line_no"),
                            "raw_label": line.get("raw_label", ""),
                            "note": line.get("note"),
                            "section": line.get("section"),
                        }
                        
                        values_json = line.get("values_json", {})
                        period_labels = line.get("period_labels", [])
                        for period, vals in values_json.items():
                            for year, value in vals.items():
                                row_data[year] = value
                        
                        all_rows.append(row_data)
                
            except Exception as e:
                print(f"  Error on page {sp.page_no}: {e}")
                import traceback
                traceback.print_exc()
        
        if all_rows:
            df = pd.DataFrame(all_rows)
            all_dfs[key] = df
            print(f"  Extracted {len(all_rows)} rows")
        else:
            print(f"  No data extracted")
    
    if not all_dfs:
        print("\nNo statement data extracted from any page.")
        return
    
    # Save all statements to Excel
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    pdf_name = pdf_path.stem
    output_file = output_dir / f"all_statements_{pdf_name}_{timestamp}.xlsx"
    
    # Clean all string columns to remove illegal Excel characters
    for key, df in all_dfs.items():
        for col in df.columns:
            if df[col].dtype == object:
                all_dfs[key][col] = df[col].apply(clean_for_excel)
    
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        for sheet_name, df in all_dfs.items():
            # Truncate sheet name if needed (Excel limit: 31 chars)
            safe_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)
            
            # Apply formatting
            worksheet = writer.sheets[safe_name]
            statement_type = sheet_name.split("_")[0]
            format_statement_sheet(worksheet, df, statement_type)
    
    print(f"\n{'='*60}")
    print(f"RESULTS SAVED TO: {output_file}")
    print(f"{'='*60}")
    
    # Print summary
    print("\n=== EXTRACTION SUMMARY ===")
    for sheet_name, df in all_dfs.items():
        print(f"\n{sheet_name}:")
        print(f"  Rows: {len(df)}")
        print(f"  Columns: {list(df.columns)}")
        
        # Show first few rows
        display_cols = [c for c in df.columns if c in ["raw_label", "note", "section"]][:3]
        value_cols = [c for c in df.columns if c not in ["page", "line_no", "raw_label", "note", "section"]][:4]
        display_cols.extend(value_cols)
        display_cols = [c for c in display_cols if c in df.columns]
        
        if display_cols:
            pd.set_option('display.max_colwidth', 40)
            print(df[display_cols].head(10).to_string())
    
    return output_file


def clear_test_results(output_dir: Path):
    """Delete all files in test_results folder."""
    if output_dir.exists():
        import shutil
        for item in output_dir.iterdir():
            try:
                if item.is_file():
                    item.unlink()
                    print(f"  Deleted: {item.name}")
                elif item.is_dir():
                    shutil.rmtree(item)
                    print(f"  Deleted folder: {item.name}")
            except PermissionError:
                print(f"  Skipped (file in use): {item.name}")
            except Exception as e:
                print(f"  Error deleting {item.name}: {e}")
    else:
        output_dir.mkdir(parents=True)


def extract_notes(pdf_path: Path, output_dir: Path):
    """Extract all notes from PDF and save to JSON."""
    from app.services.notes_store import extract_notes_structured, notes_to_json
    
    print("\n" + "=" * 60)
    print("NOTES EXTRACTION")
    print("=" * 60)
    
    pdf_bytes = pdf_path.read_bytes()
    
    # Extract GROUP notes
    print("\nExtracting GROUP notes...")
    notes = extract_notes_structured(pdf_bytes, scope="GROUP")
    print(f"  Extracted {len(notes)} notes")
    
    # Save to JSON
    json_content = notes_to_json(notes)
    json_file = output_dir / f"notes_{pdf_path.stem}.json"
    json_file.write_text(json_content, encoding="utf-8")
    print(f"  Saved to: {json_file.name}")
    
    # Save a human-readable summary
    summary_file = output_dir / f"notes_summary_{pdf_path.stem}.txt"
    with open(summary_file, "w", encoding="utf-8") as f:
        f.write("NOTES SUMMARY\n")
        f.write("=" * 60 + "\n\n")
        for note_id in sorted(notes.keys(), key=lambda x: int(x)):
            note = notes[note_id]
            f.write(f"Note {note_id}: {note.title}\n")
            f.write(f"  Pages: {note.pages}\n")
            f.write(f"  Content: {len(note.text)} chars\n")
            if note.subsections:
                for sub_id, sub in note.subsections.items():
                    f.write(f"    {sub_id}: {sub.title}\n")
            f.write("\n")
    print(f"  Summary saved to: {summary_file.name}")
    
    # Show first few notes
    print("\n  Notes found:")
    for note_id in sorted(notes.keys(), key=lambda x: int(x))[:10]:
        note = notes[note_id]
        print(f"    Note {note_id}: {note.title[:50]}... ({len(note.text)} chars)")
    if len(notes) > 10:
        print(f"    ... and {len(notes) - 10} more")
    
    return notes


def main():
    parser = argparse.ArgumentParser(description="Extract all financial statements from PDF")
    parser.add_argument("pdf", nargs="?", help="Path to PDF file")
    parser.add_argument("--no-llm", action="store_true", help="Skip LLM classification, use heuristic only")
    
    args = parser.parse_args()
    
    project_root = backend_dir.parent
    default_pdf = project_root / "shp-afs-2025.pdf"
    output_dir = project_root / "test_results"
    
    pdf_path = Path(args.pdf) if args.pdf else default_pdf
    
    print("=" * 60)
    print("FINANCIAL STATEMENTS & NOTES EXTRACTION TEST")
    print("=" * 60)
    print(f"\nPDF: {pdf_path.name}")
    print(f"Output: {output_dir}")
    
    # Clear existing test results
    print("\nClearing test_results folder...")
    clear_test_results(output_dir)
    
    # Extract all financial statements
    test_all_statements(
        pdf_path,
        use_llm=not args.no_llm
    )
    
    # Extract notes
    extract_notes(pdf_path, output_dir)
    
    print("\n" + "=" * 60)
    print("ALL TESTS COMPLETE")
    print("=" * 60)
    print(f"\nResults saved to: {output_dir}")
    for f in output_dir.iterdir():
        print(f"  - {f.name}")


if __name__ == "__main__":
    main()
